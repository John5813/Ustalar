import asyncio
import glob
import io
import logging
import os
import signal
import time
from datetime import datetime, timedelta
from dotenv import load_dotenv
from aiogram import Bot, Dispatcher
from aiogram.types import ErrorEvent, BufferedInputFile

# Load environment variables from .env file
load_dotenv()
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.fsm.storage.memory import MemoryStorage

from bot.handlers import start, documents, payments, admin, settings, samples, media
from bot.handlers import emoji as emoji_handler
from bot.handlers import converter
from bot.handlers import pptx_converter
from bot.handlers import book_translate
from bot.handlers import test as test_handler
from bot.handlers import premium_presentation as premium_presentation_handler
from bot.middlewares import LanguageMiddleware, DatabaseMiddleware
from database.database import init_db
from config import BOT_TOKEN, ADMIN_IDS
import webapp
from webapp.server import start_web_server

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def cleanup_temp_files() -> int:
    """Delete old temp files and documents.

    - temp/ images (PNG/JPG/WEBP): deleted on every call (they should not survive restarts)
    - documents/ (DOCX/PPTX/PDF): deleted only when older than 25 hours
      (tokens expire after 24h and delete their files; this catches any orphans)
    """
    removed = 0
    now = time.time()

    # Temp images — always delete on startup/periodic run
    for pattern in ["temp/*.png", "temp/*.jpg", "temp/*.jpeg", "temp/*.webp"]:
        for fp in glob.glob(pattern):
            try:
                os.remove(fp)
                removed += 1
            except Exception:
                pass

    # Generated documents — only delete orphans older than 25 hours
    doc_cutoff = now - 90000  # 25 hours
    for pattern in ["documents/*.docx", "documents/*.pptx", "documents/*.pdf"]:
        for fp in glob.glob(pattern):
            try:
                if os.path.getmtime(fp) < doc_cutoff:
                    os.remove(fp)
                    removed += 1
            except Exception:
                pass

    if removed:
        logger.info(f"Temp cleanup: removed {removed} old file(s)")
    return removed


async def generate_daily_excel(date_str: str) -> bytes:
    """Generate Excel file of users who registered on date_str (YYYY-MM-DD)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from database.database import DATABASE_FILE
    import aiosqlite

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Yangi foydalanuvchilar"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E86C1")
    headers = ["№", "Telegram ID", "Ism", "Username", "Til", "Balans (so'm)", "Ro'yxatdan o'tgan vaqt"]
    col_widths = [5, 15, 20, 20, 6, 15, 22]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.row_dimensions[1].height = 20

    # Fetch users
    async with aiosqlite.connect(DATABASE_FILE) as db_conn:
        db_conn.row_factory = aiosqlite.Row
        async with db_conn.execute(
            "SELECT telegram_id, first_name, username, language, balance, created_at "
            "FROM users WHERE date(created_at) = ? ORDER BY created_at",
            (date_str,)
        ) as cursor:
            rows = await cursor.fetchall()

    even_fill = PatternFill("solid", fgColor="EBF5FB")
    for row_idx, row in enumerate(rows, 2):
        username = f"@{row['username']}" if row['username'] else "—"
        created = row['created_at'] or ""
        # Trim microseconds if present
        if "." in created:
            created = created[:19]
        values = [
            row_idx - 1,
            row['telegram_id'],
            row['first_name'] or "—",
            username,
            (row['language'] or "uz").upper(),
            row['balance'] or 0,
            created,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if row_idx % 2 == 0:
                cell.fill = even_fill

    # Summary row
    summary_row = len(rows) + 2
    ws.cell(row=summary_row, column=1, value="Jami:").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=len(rows)).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def daily_user_report(bot: Bot):
    """Every day at 23:55 Uzbekistan time (18:55 UTC) send new-user Excel to admins."""
    UZT_OFFSET = timedelta(hours=5)
    REPORT_HOUR = 23
    REPORT_MINUTE = 55

    while True:
        # Current time in Uzbekistan (UTC+5)
        now_uzt = datetime.utcnow() + UZT_OFFSET
        target = now_uzt.replace(hour=REPORT_HOUR, minute=REPORT_MINUTE, second=0, microsecond=0)
        if now_uzt >= target:
            # Already past today's target — schedule for tomorrow
            target += timedelta(days=1)
        wait_seconds = (target - now_uzt).total_seconds()
        logger.info(f"Daily report scheduled in {wait_seconds/3600:.1f}h (at {target.strftime('%H:%M')} UZT)")
        await asyncio.sleep(wait_seconds)

        # Generate report for today
        report_date = (datetime.utcnow() + UZT_OFFSET).strftime("%Y-%m-%d")
        display_date = (datetime.utcnow() + UZT_OFFSET).strftime("%d.%m.%Y")
        try:
            excel_bytes = await generate_daily_excel(report_date)
        except Exception as exc:
            logger.error(f"Daily Excel generation failed: {exc}", exc_info=True)
            await asyncio.sleep(60)
            continue

        filename = f"yangi_foydalanuvchilar_{report_date}.xlsx"
        caption = (
            f"📊 <b>Kunlik hisobot — {display_date}</b>\n\n"
            f"Bugun ro'yxatdan o'tgan foydalanuvchilar ro'yxati."
        )
        from config import ADMIN_IDS
        for admin_id in ADMIN_IDS:
            try:
                await bot.send_document(
                    chat_id=admin_id,
                    document=BufferedInputFile(excel_bytes, filename=filename),
                    caption=caption,
                    parse_mode="HTML",
                )
            except Exception as send_exc:
                logger.warning(f"Could not send daily report to admin {admin_id}: {send_exc}")

        # Wait a bit to avoid double-send if sleep wakes slightly early
        await asyncio.sleep(120)


async def periodic_cleanup(interval_seconds: int = 1800, storage=None):
    """Run cleanup every interval_seconds (default 30 min).

    Note: The FSM storage is NOT periodically wiped anymore. The previous
    behaviour cleared every active user's state every 4 hours, which abruptly
    interrupted users mid-conversation. /start already calls state.clear(),
    which is enough to recover any stuck user.
    """
    _ = storage  # kept for signature compatibility
    while True:
        await asyncio.sleep(interval_seconds)
        cleanup_temp_files()
        # Release matplotlib global figure registry
        try:
            import matplotlib.pyplot as plt
            plt.close("all")
        except Exception:
            pass
        # Force Python GC
        try:
            import gc
            gc.collect()
        except Exception:
            pass
        # Return freed memory pages back to OS (Linux glibc only)
        try:
            import ctypes
            ctypes.cdll.LoadLibrary("libc.so.6").malloc_trim(0)
        except Exception:
            pass
        logger.info("Periodic memory cleanup done (plt.close('all') + gc.collect())")


async def main():
    """Main function to start the bot"""
    # Initialize database
    await init_db()

    # Set Mini App domain from environment
    webapp.WEBAPP_DOMAIN = os.environ.get("REPLIT_DEV_DOMAIN", "localhost:5000")

    # Restore tokens saved before last restart
    webapp.load_tokens_from_disk()

    # Initialize bot and dispatcher
    bot = Bot(
        token=BOT_TOKEN,
        default=DefaultBotProperties(parse_mode=ParseMode.HTML)
    )
    
    dp = Dispatcher(storage=MemoryStorage())

    # ── Global error handler ───────────────────────────────────────────────────
    # Catches TelegramBadRequest (expired/invalid callback query IDs) and other
    # Telegram API errors BEFORE they bubble up through the dispatcher and create
    # a tight CPU-burning retry loop.  Without this, every stale inline-button
    # click causes an unhandled exception that aiogram keeps re-propagating,
    # pegging the CPU at ~100%.
    @dp.errors()
    async def global_error_handler(event: ErrorEvent) -> bool:
        exception = event.exception
        exc_name = type(exception).__name__
        exc_msg = str(exception)

        # Silently drop expired / invalid callback query errors.
        # These happen when a user clicks an inline button that is older than
        # Telegram's 48-hour callback-query lifetime.
        stale_callback_phrases = (
            "query is too old",
            "query ID is invalid",
            "MESSAGE_ID_INVALID",
            "message to edit not found",
            "message is not modified",
            "message can't be edited",
        )
        if any(phrase in exc_msg for phrase in stale_callback_phrases):
            logger.debug(f"Ignored stale callback error [{exc_name}]: {exc_msg[:120]}")
            # Try to silently answer the callback so Telegram stops showing
            # the loading spinner on the user's device.
            update = event.update
            if update and update.callback_query:
                try:
                    await update.callback_query.answer()
                except Exception:
                    pass
            return True  # mark as handled — do NOT re-raise

        # Drop "bot was blocked / kicked" errors silently (common with broadcasts)
        blocked_phrases = (
            "bot was blocked by the user",
            "user is deactivated",
            "chat not found",
            "bot was kicked",
            "Forbidden",
        )
        if any(phrase in exc_msg for phrase in blocked_phrases):
            logger.debug(f"Ignored blocked-user error [{exc_name}]: {exc_msg[:120]}")
            return True

        # Drop network / flood-control errors that resolve on their own
        transient_phrases = (
            "Too Many Requests",
            "retry_after",
            "FLOOD_WAIT",
            "Connection",
            "TimeoutError",
            "ServerDisconnectedError",
        )
        if any(phrase in exc_msg for phrase in transient_phrases):
            logger.warning(f"Transient Telegram error [{exc_name}]: {exc_msg[:120]}")
            return True

        # All other errors: log them with full traceback but do NOT crash the bot
        logger.error(
            f"Unhandled error in update handler [{exc_name}]: {exc_msg[:300]}",
            exc_info=exception,
        )
        return True  # returning True prevents aiogram from re-raising

    # Register middlewares
    dp.message.middleware(DatabaseMiddleware())
    dp.callback_query.middleware(DatabaseMiddleware())
    dp.pre_checkout_query.middleware(DatabaseMiddleware())
    dp.message.middleware(LanguageMiddleware())
    dp.callback_query.middleware(LanguageMiddleware())
    dp.pre_checkout_query.middleware(LanguageMiddleware())
    
    # Block check middleware - must be last to check after database is injected
    from bot.middlewares import BlockedUserMiddleware
    dp.message.middleware(BlockedUserMiddleware())
    dp.callback_query.middleware(BlockedUserMiddleware())
    
    # Register handlers - important order: specific handlers first, catch-all last!
    dp.include_router(admin.router)  # Admin commands first
    dp.include_router(settings.router)  # Handle settings buttons
    dp.include_router(converter.router)  # Handle PDF → DOCX conversion (before payments to keep state-specific callbacks)
    dp.include_router(pptx_converter.router)  # Handle PPTX → PDF conversion
    dp.include_router(payments.router)  # Handle payment buttons
    dp.include_router(samples.router)  # Handle samples view and admin management
    dp.include_router(emoji_handler.router)  # Handle emoji mosaic conversion
    dp.include_router(media.router)   # Legacy media router (empty)
    dp.include_router(book_translate.router)  # Handle book translation service
    dp.include_router(test_handler.router)  # Handle test generation service
    dp.include_router(premium_presentation_handler.router)  # Premium taqdimot — Ustalar tizimi
    dp.include_router(documents.router)  # Handles document creation and topic input - MUST BE BEFORE start.router
    dp.include_router(start.router)  # LAST - has catch-all handler for unknown messages
    
    # Share bot instance with web server for sending files
    webapp.BOT = bot

    # Clean up old temp files left from previous runs
    cleanup_temp_files()

    # Delete any existing webhook before starting polling
    # drop_pending_updates=True ensures stale callback queries accumulated
    # while the bot was offline do NOT flood the handler on startup.
    await bot.delete_webhook(drop_pending_updates=True)

    # Fetch bot's own username and pass it to the sticker handler
    bot_info = await bot.get_me()
    emoji_handler.set_bot_username(bot_info.username or "")
    
    # Start the heavy-document generation queue (single worker).
    # Heavy docs (kurs ishi, diplom ishi, dissertatsiya, bitiruv ishi) are
    # serialised here so only ONE big generation runs at a time, preventing
    # OOM spikes on the production server.
    from bot.queue_service import get_doc_queue
    get_doc_queue().start()

    # Start both bot and web server concurrently
    logger.info("Bot started")
    logger.info(f"Mini App domain: {webapp.WEBAPP_DOMAIN}")

    loop = asyncio.get_running_loop()
    stop_event = asyncio.Event()

    def _signal_handler():
        logger.info("Shutdown signal received, stopping bot...")
        stop_event.set()

    for sig in (signal.SIGINT, signal.SIGTERM):
        try:
            loop.add_signal_handler(sig, _signal_handler)
        except NotImplementedError:
            pass

    polling_task = asyncio.create_task(dp.start_polling(bot))
    web_task     = asyncio.create_task(start_web_server(port=5000))
    cleanup_task = asyncio.create_task(periodic_cleanup(storage=dp.storage))
    report_task  = asyncio.create_task(daily_user_report(bot))

    try:
        await stop_event.wait()
    finally:
        logger.info("Cancelling tasks...")
        for task in (polling_task, web_task, cleanup_task, report_task):
            task.cancel()
        await asyncio.gather(polling_task, web_task, cleanup_task, report_task, return_exceptions=True)

        await bot.session.close()
        from services.ai_service import close_ai_service
        from services.together_service import close_together_service
        await close_ai_service()
        await close_together_service()
        logger.info("Bot stopped cleanly")

if __name__ == "__main__":
    asyncio.run(main())
